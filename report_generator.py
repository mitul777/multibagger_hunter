# =============================================================================
# report_generator.py — Output Results as Console Table, Excel & HTML
# =============================================================================

import os
import math
import numpy as np
import pandas as pd
from datetime import datetime
from rich.console import Console
from rich.table import Table
from rich import box
from config import OUTPUT, SCORE_WEIGHTS

console = Console()


class ReportGenerator:

    def __init__(self, output_dir: str = "output"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M")

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def generate(self, results: list[dict]) -> None:
        """Generate all report formats."""
        df = self._to_dataframe(results)

        if df.empty:
            console.print("[red]No results to display.[/red]")
            return

        top_n = OUTPUT.get("top_n_display", 25)

        self._print_console(df, top_n)
        self._print_summary_stats(df)

        if OUTPUT.get("excel_output"):
            self._save_excel(df)

        if OUTPUT.get("html_output"):
            self._save_html(df, top_n)

        console.print(f"\n[green]✓ Reports saved to:[/green] [bold]{self.output_dir}/[/bold]")

    # ------------------------------------------------------------------
    # Console Table
    # ------------------------------------------------------------------

    def _print_console(self, df: pd.DataFrame, top_n: int) -> None:
        top = df.head(top_n)

        table = Table(
            title=f"\n🏆 Top {top_n} Indian Equity Moat Candidates",
            box=box.ROUNDED,
            show_lines=True,
            header_style="bold cyan",
        )

        columns = [
            ("Rank",       "right",  4),
            ("Ticker",     "left",   12),
            ("Company",    "left",   28),
            ("Sector",     "left",   22),
            ("Score",      "right",  7),
            ("Moat Label", "left",   18),
            ("ROCE%",      "right",  8),
            ("Gr.Margin%", "right",  10),
            ("Rev CAGR%",  "right",  10),
            ("D/E",        "right",  6),
            ("CFO/PAT",    "right",  8),
            ("Key Signals","left",   40),
        ]

        for name, justify, _ in columns:
            table.add_column(name, justify=justify, no_wrap=(name == "Ticker"))

        for rank, (_, row) in enumerate(top.iterrows(), 1):
            signals = "; ".join(row.get("moat_signals", [])[:2])  # top 2 signals
            table.add_row(
                str(rank),
                str(row.get("ticker", "")),
                self._truncate(str(row.get("name", "")), 27),
                self._truncate(str(row.get("sector", "")), 21),
                f"[bold yellow]{row['score_total']:.1f}[/bold yellow]",
                str(row.get("moat_label", "")),
                self._fmt_pct(row.get("roce_avg")),
                self._fmt_pct(row.get("gross_margin_avg")),
                self._fmt_pct(row.get("rev_cagr")),
                self._fmt_ratio(row.get("debt_equity_latest")),
                self._fmt_ratio(row.get("cfo_pat_avg")),
                signals,
            )

        console.print(table)

    # ------------------------------------------------------------------
    # Summary Statistics
    # ------------------------------------------------------------------

    def _print_summary_stats(self, df: pd.DataFrame) -> None:
        console.print("\n[bold cyan]━━━  SCREENING SUMMARY  ━━━[/bold cyan]")
        console.print(f"  Total companies screened   : {len(df)}")

        thresholds = OUTPUT["score_thresholds"]
        strong  = (df["score_total"] >= thresholds["strong_moat"]).sum()
        emerging = (df["score_total"] >= thresholds["emerging_moat"]).sum() - strong
        watch   = (df["score_total"] >= thresholds["watchlist"]).sum() - strong - emerging
        no_moat = len(df) - strong - emerging - watch

        console.print(f"  🏰 Strong Moat  (≥{thresholds['strong_moat']}pts)  : {strong}")
        console.print(f"  🌱 Emerging Moat(≥{thresholds['emerging_moat']}pts)  : {emerging}")
        console.print(f"  👀 Watchlist    (≥{thresholds['watchlist']}pts)  : {watch}")
        console.print(f"  ❌ No Moat      (<{thresholds['watchlist']}pts)   : {no_moat}")

        console.print("\n[bold cyan]━━━  SCORE BREAKDOWN AVERAGES (Top 25) ━━━[/bold cyan]")
        top25 = df.head(25)
        for col, label in [
            ("score_moat",       f"Moat      (max {SCORE_WEIGHTS['moat']}pts)"),
            ("score_quality",    f"Quality   (max {SCORE_WEIGHTS['quality']}pts)"),
            ("score_growth",     f"Growth    (max {SCORE_WEIGHTS['growth']}pts)"),
            ("score_efficiency", f"Efficiency(max {SCORE_WEIGHTS['efficiency']}pts)"),
        ]:
            avg = top25[col].mean() if col in top25.columns else 0
            console.print(f"  {label}: avg = {avg:.1f}")

    # ------------------------------------------------------------------
    # Excel Export
    # ------------------------------------------------------------------

    def _save_excel(self, df: pd.DataFrame) -> None:
        path = os.path.join(self.output_dir, f"moat_screen_{self.timestamp}.xlsx")

        display_cols = [
            "ticker", "name", "sector", "industry",
            "score_total", "score_moat", "score_quality", "score_growth",
            "score_efficiency", "moat_label",
            "market_cap_cr", "current_price",
            "roce_avg", "roce_latest", "roce_cv",
            "gross_margin_avg", "gross_margin_latest", "gross_margin_cv",
            "op_margin_avg", "op_margin_latest", "op_margin_trend",
            "net_margin_latest",
            "rev_cagr", "rev_cagr_3y", "rev_cagr_5y",
            "eps_cagr", "ni_cagr_3y", "ni_cagr_5y",
            "roe_avg", "roe_latest",
            "debt_equity_latest", "debt_equity_avg",
            "cfo_pat_avg", "fcf_conversion_avg", "fcf_latest",
            "interest_coverage_avg", "net_cash_cr",
            "asset_turnover_avg", "operating_leverage",
            "current_ratio",
            "pe_ratio", "pb_ratio", "ev_ebitda", "peg_ratio",
        ]

        # Only keep columns that exist
        cols = [c for c in display_cols if c in df.columns]
        export_df = df[cols].copy()

        # Round numeric columns
        for col in export_df.select_dtypes(include=[np.number]).columns:
            export_df[col] = export_df[col].round(2)

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            export_df.to_excel(writer, sheet_name="Moat Screener", index=False)
            self._format_excel(writer, export_df)

        console.print(f"  📊 Excel  → {path}")

    def _format_excel(self, writer, df: pd.DataFrame) -> None:
        """Apply basic formatting to Excel output."""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            ws = writer.sheets["Moat Screener"]

            # Header formatting
            header_fill = PatternFill("solid", fgColor="1F4E79")
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Freeze header row
            ws.freeze_panes = "A2"

            # Auto-width (approximate)
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

            # Color rows by moat label
            moat_col = None
            for i, cell in enumerate(ws[1], 1):
                if cell.value == "moat_label":
                    moat_col = i
                    break

            if moat_col:
                for row in ws.iter_rows(min_row=2):
                    label_cell = row[moat_col - 1]
                    if label_cell.value and "Strong" in str(label_cell.value):
                        for cell in row:
                            cell.fill = PatternFill("solid", fgColor="C6EFCE")
                    elif label_cell.value and "Emerging" in str(label_cell.value):
                        for cell in row:
                            cell.fill = PatternFill("solid", fgColor="FFEB9C")
        except Exception:
            pass  # Formatting is optional

    # ------------------------------------------------------------------
    # HTML Export
    # ------------------------------------------------------------------

    def _save_html(self, df: pd.DataFrame, top_n: int) -> None:
        path = os.path.join(self.output_dir, f"moat_screen_{self.timestamp}.html")
        top = df.head(top_n)

        rows_html = ""
        for rank, (_, row) in enumerate(top.iterrows(), 1):
            label = row.get("moat_label", "")
            row_class = (
                "strong" if "Strong" in str(label) else
                "emerging" if "Emerging" in str(label) else
                "watchlist" if "Watch" in str(label) else "none"
            )
            signals = "<br>".join(row.get("moat_signals", [])[:4])

            rows_html += f"""
            <tr class="{row_class}">
                <td><b>{rank}</b></td>
                <td><code>{row.get('ticker','')}</code></td>
                <td>{self._truncate(str(row.get('name','')), 35)}</td>
                <td>{self._truncate(str(row.get('sector','')), 25)}</td>
                <td><b>{row['score_total']:.1f}</b></td>
                <td>{label}</td>
                <td>{self._fmt_pct(row.get('roce_avg'))}</td>
                <td>{self._fmt_pct(row.get('gross_margin_avg'))}</td>
                <td>{self._fmt_pct(row.get('rev_cagr'))}</td>
                <td>{self._fmt_ratio(row.get('debt_equity_latest'))}</td>
                <td>{self._fmt_ratio(row.get('cfo_pat_avg'))}</td>
                <td class="signals">{signals}</td>
            </tr>"""

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Indian Equity Moat Screener — {self.timestamp}</title>
<style>
  body {{ font-family: 'Segoe UI', sans-serif; background: #0d1117; color: #c9d1d9; margin: 20px; }}
  h1   {{ color: #58a6ff; border-bottom: 2px solid #30363d; padding-bottom: 10px; }}
  h2   {{ color: #79c0ff; }}
  .meta {{ color: #8b949e; font-size: 0.85em; margin-bottom: 20px; }}
  table {{ border-collapse: collapse; width: 100%; font-size: 0.85em; }}
  th    {{ background: #1f6feb; color: #fff; padding: 8px 10px; text-align: left; position: sticky; top: 0; }}
  td    {{ padding: 6px 10px; border-bottom: 1px solid #21262d; vertical-align: top; }}
  tr.strong   {{ background: #0d2818; }}
  tr.emerging {{ background: #1c1a0a; }}
  tr.watchlist{{ background: #0a1a2e; }}
  tr:hover    {{ background: #161b22 !important; }}
  code   {{ background: #21262d; padding: 2px 6px; border-radius: 4px; color: #79c0ff; }}
  .signals {{ font-size: 0.80em; color: #8b949e; max-width: 280px; }}
  .badge {{ display: inline-block; padding: 2px 8px; border-radius: 10px; font-size: 0.80em; font-weight: bold; }}
  .legend {{ display: flex; gap: 20px; margin: 15px 0; flex-wrap: wrap; }}
  .legend-item {{ display: flex; align-items: center; gap: 8px; }}
  .dot {{ width: 14px; height: 14px; border-radius: 50%; }}
  .dot-strong   {{ background: #3fb950; }}
  .dot-emerging {{ background: #d29922; }}
  .dot-watch    {{ background: #388bfd; }}
</style>
</head>
<body>
<h1>🏰 Indian Equity — Moat Screener Results</h1>
<div class="meta">
  Generated: {datetime.now().strftime("%d %b %Y, %H:%M")} &nbsp;|&nbsp;
  Powered by: yfinance &nbsp;|&nbsp;
  Universe: NSE-listed Nifty 500 companies &nbsp;|&nbsp;
  Total screened: {len(df)} stocks
</div>

<div class="legend">
  <div class="legend-item"><div class="dot dot-strong"></div> Strong Moat (≥70 pts)</div>
  <div class="legend-item"><div class="dot dot-emerging"></div> Emerging Moat (≥50 pts)</div>
  <div class="legend-item"><div class="dot dot-watch"></div> Watchlist (≥35 pts)</div>
</div>

<h2>Top {top_n} Companies by Moat Score</h2>
<table>
  <thead>
    <tr>
      <th>#</th><th>Ticker</th><th>Company</th><th>Sector</th>
      <th>Score /100</th><th>Moat Label</th>
      <th>ROCE%</th><th>Gross Margin%</th><th>Rev CAGR%</th>
      <th>D/E</th><th>CFO/PAT</th><th>Key Signals</th>
    </tr>
  </thead>
  <tbody>
    {rows_html}
  </tbody>
</table>

<br>
<h2>📊 Scoring Methodology</h2>
<table style="max-width: 700px;">
  <thead><tr><th>Dimension</th><th>Weight</th><th>Key Metrics</th></tr></thead>
  <tbody>
    <tr><td>🏰 Moat</td><td>40 pts</td><td>ROCE (avg & consistency), Gross Margins, FCF Conversion</td></tr>
    <tr><td>💎 Quality</td><td>25 pts</td><td>Debt/Equity, CFO/PAT ratio, Interest Coverage</td></tr>
    <tr><td>📈 Growth</td><td>25 pts</td><td>Revenue CAGR, EPS/Net Income CAGR (3yr & 5yr)</td></tr>
    <tr><td>⚙️ Efficiency</td><td>10 pts</td><td>Asset Turnover, Operating Leverage</td></tr>
  </tbody>
</table>

<div class="meta" style="margin-top: 30px;">
  ⚠️ This tool is for educational/research purposes only. Not financial advice. Always do your own due diligence.
</div>
</body>
</html>"""

        with open(path, "w") as f:
            f.write(html)
        console.print(f"  🌐 HTML   → {path}")

    # ------------------------------------------------------------------
    # Utilities
    # ------------------------------------------------------------------

    def _to_dataframe(self, results: list[dict]) -> pd.DataFrame:
        if not results:
            return pd.DataFrame()
        df = pd.DataFrame(results)
        df = df.sort_values("score_total", ascending=False).reset_index(drop=True)
        # Convert moat_signals list to list (keep for HTML; join for display)
        return df

    @staticmethod
    def _fmt_pct(val) -> str:
        if val is None or (isinstance(val, float) and math.isnan(val)):
            return "—"
        return f"{val:.1f}%"

    @staticmethod
    def _fmt_ratio(val) -> str:
        if val is None or (isinstance(val, float) and math.isnan(val)):
            return "—"
        return f"{val:.2f}x"

    @staticmethod
    def _truncate(s: str, n: int) -> str:
        return s[:n] + "…" if len(s) > n else s
